from openpyxl import load_workbook
from file_downloader import FileDownloader
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from cStringIO import StringIO
from elastic_search import ElasticSearch
from entity.GBP import GBP
import json

class DataLoader(object):
    def __init__(self, file_dir):
        self.file_dir = file_dir
        self.es = ElasticSearch()
        status = self.es.check_node_status()
        if status != None:
            self.es.connect_es()

    def read_excel_file(self):
        workbook = load_workbook(self.file_dir + 'GBP-SBP-resource-centre-250917.xlsx')
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)

        for i, row in enumerate(worksheet.iter_rows()):
            if i == 0:
                continue
            if self.es.check_document_exists(index="green_bond", doc="report", id=i):
                print "Document " + str(i) + " exists!"
                doc = self.es.find_document(index="green_bond", doc="report", id=i)
                if row[4].hyperlink is not None and row[4].internal_value == doc["_source"]["external_review_report"]:
                    try:
                        link = row[4].hyperlink.target
                        file_downloader = FileDownloader()
                        file_downloader.download_file(link)
                        text = self.read_pdf_file(self.file_dir + "temp_file.pdf")
                        gbp = GBP(row[0].internal_value, row[1].internal_value, row[2].internal_value,
                                  row[3].internal_value, text, row[5].internal_value)
                        self.es.update_document(index="green_bond", doc="report", id=i, data=json.dumps(gbp.__dict__))
                        print "Update document " + str(i) + " succesfully!"
                    except:
                        print "Hyperlink error! Document " + str(i)
                        pass
                if row[3].hyperlink is not None and row[3].internal_value == doc["_source"]["external_review_form"]:
                    try:
                        link = row[3].hyperlink.target
                        file_downloader = FileDownloader()
                        file_downloader.download_file(link)
                        text = self.read_pdf_file(self.file_dir + "temp_file.pdf")
                        gbp = GBP(doc["_source"]["green_bond_issuer"], doc["_source"]["country"], doc["_source"]["market_information_template"],
                                  text, doc["_source"]["external_review_report"], doc["_source"]["external_link"])
                        self.es.update_document(index="green_bond", doc="report", id=i, data=json.dumps(gbp.__dict__))
                        print "Update document " + str(i) + " succesfully!"
                    except:
                        print "Hyperlink error! Document " + str(i)
                        pass
                continue

            if row[4].hyperlink is None:
                gbp = GBP(row[0].internal_value, row[1].internal_value, row[2].internal_value,
                          row[3].internal_value, row[4].internal_value, row[5].internal_value)
                self.es.add_document(i, json.dumps(gbp.__dict__))
                print "Adding document " + str(i) + " succesfully!"

            else:
                try:
                    link = row[4].hyperlink.target
                    file_downloader = FileDownloader()
                    file_downloader.download_file(link)
                    self.text = self.read_pdf_file(self.file_dir + "temp_file.pdf")
                    gbp = GBP(row[0].internal_value, row[1].internal_value, row[2].internal_value,
                          row[3].internal_value, self.text, row[5].internal_value)
                    self.es.add_document(i, json.dumps(gbp.__dict__))
                    print "Adding document " + str(i) + " succesfully!"
                except:
                    print "Hyperlink error! Document " + str(i)
                    pass

    def read_pdf_file(self, pdf_file):
        rsrcmgr = PDFResourceManager()
        retstr = StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = file(pdf_file, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 0
        caching = True
        pagenos = set()

        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password, caching=caching, check_extractable=True):
            interpreter.process_page(page)

        text = retstr.getvalue()
        fp.close()
        device.close()
        retstr.close()
        return text
